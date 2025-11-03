#!/usr/bin/env python3
"""
NANPA Prefix Updater
--------------------
Downloads the latest NANPA CO Code Assignment data (Available + Utilized),
extracts all .txt/.csv/.xlsx files, parses and normalizes prefixes,
and saves a unified data.json + carriers.db with proper carrier/type fields.

This version:
- Uses parent-only branding for `carrier` (e.g., AT&T, Verizon, T-Mobile, UScellular, DISH)
- Never guesses type: leaves `type` blank unless confidently Mobile / VoIP / Paging / Landline
- Adds `company_original` to preserve the original NANPA company string

Usage:
  python nanpa_updater.py                 # full update (scrape nanpa.com, download zips, extract, parse, save)
  python nanpa_updater.py --check-only    # list .zip links found + Last-Modified headers (no download/parse)
  python nanpa_updater.py --no-fetch      # skip download, reprocess existing local zips in ./nanpa_zips

Requires: requests, openpyxl (install via pip)
"""
import os, json, csv, sys, zipfile, shutil, sqlite3, datetime, argparse
from urllib.parse import urljoin

import requests
from requests.adapters import HTTPAdapter, Retry
from openpyxl import load_workbook

# ---------------------------- CONFIG ---------------------------------
INDEX_URL = "https://www.nanpa.com/reports/co-code-reports/cocodes_assign"  # page that lists the ZIPs
OUTPUT_JSON = "data.json"
OUTPUT_DB = "carriers.db"
FILES_DIR = "nanpa_files"
BACKUP_DIR = "backups"
ZIPS_DIR = "nanpa_zips"

STATE_MAP = {
 "AL":"Alabama","AK":"Alaska","AZ":"Arizona","AR":"Arkansas","CA":"California","CO":"Colorado","CT":"Connecticut","DE":"Delaware","DC":"District of Columbia","FL":"Florida","GA":"Georgia","HI":"Hawaii","ID":"Idaho","IL":"Illinois","IN":"Indiana","IA":"Iowa","KS":"Kansas","KY":"Kentucky","LA":"Louisiana","ME":"Maine","MD":"Maryland","MA":"Massachusetts","MI":"Michigan","MN":"Minnesota","MS":"Mississippi","MO":"Missouri","MT":"Montana","NE":"Nebraska","NV":"Nevada","NH":"New Hampshire","NJ":"New Jersey","NM":"New Mexico","NY":"New York","NC":"North Carolina","ND":"North Dakota","OH":"Ohio","OK":"Oklahoma","OR":"Oregon","PA":"Pennsylvania","RI":"Rhode Island","SC":"South Carolina","SD":"South Dakota","TN":"Tennessee","TX":"Texas","UT":"Utah","VT":"Vermont","VA":"Virginia","WA":"Washington","WV":"West Virginia","WI":"Wisconsin","WY":"Wyoming"
}

# ----------- BRAND & TYPE CLASSIFICATION (no regex, substring-based) -----------
# Parent-only brand mapping
BRAND_KEYWORDS = [
    # AT&T umbrella: AT&T Mobility, New Cingular, Cricket, FirstNet
    (["NEW CINGULAR", "AT&T MOBILITY", "AT&T", "CRICKET", "FIRSTNET", "FIRST NET"], "AT&T"),
    # Verizon umbrella: Cellco Partnership, Verizon Wireless, TracFone & brands
    (["CELLCO PARTNERSHIP", "VERIZON WIRELESS", "VERIZON", "TRACFONE", "STRAIGHT TALK", "TOTAL BY VERIZON", "SIMPLE MOBILE", "NET10", "NET 10", "PAGE PLUS", "SAFELINK"], "Verizon"),
    # T-Mobile umbrella: T-Mobile, Metro/MetroPCS, Sprint Spectrum, Clearwire, Mint, Ultra Mobile, Google Fi
    (["T-MOBILE", "TMOBILE", "METROPCS", "METRO BY T-MOBILE", "METRO ", "SPRINT SPECTRUM", "SPRINT", "CLEARWIRE", "MINT MOBILE", "ULTRA MOBILE", "GOOGLE FI"], "T-Mobile"),
    # UScellular
    (["US CELLULAR", "UNITED STATES CELLULAR"], "UScellular"),
    # DISH / Boost
    (["DISH WIRELESS", "BOOST MOBILE", "BOOST"], "DISH"),
]

# Type keyword families (substring match, uppercased)
WIRELESS_KEYWORDS = ["WIRELESS", "MOBILE", "CELL", "PCS", "LTE", "5G"]
VOIP_KEYWORDS     = ["VOIP", "VONAGE", "BANDWIDTH", "LEVEL 3", "L3", "LUMEN VOIP", "TWILIO", "ZOOM PHONE"]
PAGING_KEYWORDS   = ["PAGING", "USA MOBILITY", "AMERICAN MESSAGING"]
# Landline identification (only mark Landline if one of these fixed-line brands appears)
LANDLINE_KEYWORDS = [
    "FRONTIER", "WINDSTREAM", "CONSOLIDATED COMMUNICATIONS",
    "CENTURYLINK", "EMBARQ", "QWEST", "LUMEN",  # Lumen family (non-Level 3 VOIP)
    "COX", "COMCAST", "XFINITY", "CHARTER", "SPECTRUM",
    # Landline arms of AT&T/Verizon when not wireless
    "VERIZON", "AT&T"
]

# ---------------------------- HELPERS ---------------------------------
def log(msg: str):
    print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {msg}")

def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) NANPA-Updater/1.2",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    })
    retries = Retry(total=5, backoff_factor=0.5, status_forcelist=[429, 500, 502, 503, 504], allowed_methods=["GET","HEAD"])
    s.mount("https://", HTTPAdapter(max_retries=retries))
    s.mount("http://", HTTPAdapter(max_retries=retries))
    return s

def normalize_brand(company: str) -> str:
    c = (company or "").upper()
    for keywords, brand in BRAND_KEYWORDS:
        for kw in keywords:
            if kw in c:
                return brand
    return ""  # unknown parent

def detect_type(company: str, brand: str) -> str:
    """Return one of: 'Mobile','VoIP','Paging','Landline','' (blank if unknown)."""
    c = (company or "").upper()
    # Mobile if brand is a known wireless parent
    if brand in {"AT&T", "Verizon", "T-Mobile", "UScellular", "DISH"}:
        return "Mobile"
    # Or if explicit wireless keywords
    if any(kw in c for kw in WIRELESS_KEYWORDS):
        return "Mobile"
    # VoIP vendors
    if any(kw in c for kw in VOIP_KEYWORDS):
        return "VoIP"
    # Paging vendors
    if any(kw in c for kw in PAGING_KEYWORDS):
        return "Paging"
    # Landline only when clearly matched to known fixed-line brands
    if any(kw in c for kw in LANDLINE_KEYWORDS):
        # Avoid mislabeling VOIP as Landline if Level 3/Bandwidth/Twilio matched above
        return "Landline"
    return ""  # unknown

def safe_title(s: str) -> str:
    return s.title().replace("_", " ") if s else ""

# ---------------------------- SCRAPE ZIP LINKS ---------------------------------
def get_zip_links(session: requests.Session) -> list:
    log(f"Scraping {INDEX_URL} for .zip links ...")
    resp = session.get(INDEX_URL, timeout=30)
    resp.raise_for_status()
    html = resp.text
    # naive find of hrefs ending with .zip (absolute or relative)
    links = []
    start = 0
    while True:
        i = html.find(".zip", start)
        if i == -1:
            break
        # backtrack to find href="..."
        j = html.rfind("href=\"", 0, i)
        if j == -1:
            start = i + 4
            continue
        j += len("href=\"")
        k = html.find('"', j)
        if k == -1:
            start = i + 4
            continue
        href = html[j:k]
        if href.lower().endswith('.zip'):
            links.append(urljoin(INDEX_URL, href))
        start = i + 4
    # de-dup while preserving order
    seen, uniq = set(), []
    for u in links:
        if u not in seen:
            seen.add(u)
            uniq.append(u)
    if not uniq:
        log("WARN: No .zip links found on the page. The page structure may have changed.")
    else:
        log(f"Found {len(uniq)} zip link(s).")
    return uniq

# ---------------------------- FETCH ---------------------------------
def fetch_latest(zip_dir: str, session: requests.Session) -> list:
    os.makedirs(zip_dir, exist_ok=True)
    new_files = []
    links = get_zip_links(session)
    for url in links:
        fname = os.path.join(zip_dir, os.path.basename(url))
        try:
            # Try to skip if same size as remote
            remote_size = 0
            try:
                head = session.head(url, allow_redirects=True, timeout=20)
                remote_size = int(head.headers.get("Content-Length", "0") or 0)
            except Exception:
                pass

            if os.path.exists(fname) and remote_size and os.path.getsize(fname) == remote_size:
                log(f"  Skipped (already latest): {os.path.basename(fname)}")
                continue

            log(f"  Downloading {url}")
            r = session.get(url, timeout=60)
            r.raise_for_status()
            with open(fname, "wb") as f:
                f.write(r.content)
            new_files.append(fname)
        except Exception as e:
            log(f"  WARN: failed {url} -> {e}")
    return new_files

# ---------------------------- EXTRACT ---------------------------------
def extract_all(zip_dir: str, out_dir: str):
    # Clear output to avoid mixing old/new schema
    if os.path.exists(out_dir):
        shutil.rmtree(out_dir)
    os.makedirs(out_dir, exist_ok=True)

    found = False
    for z in os.listdir(zip_dir):
        if not z.lower().endswith('.zip'):
            continue
        zpath = os.path.join(zip_dir, z)
        try:
            with zipfile.ZipFile(zpath, 'r') as zip_ref:
                zip_ref.extractall(out_dir)
                log(f"Extracted {z} -> {out_dir}")
                found = True
        except zipfile.BadZipFile:
            log(f"WARN: Corrupt zip skipped: {z}")
    if not found:
        log("WARN: No zip files extracted (nothing to parse).")

# ---------------------------- PARSE ---------------------------------
def iter_rows_from_file(fpath: str):
    ext = os.path.splitext(fpath)[1].lower()
    if ext == '.xlsx':
        wb = load_workbook(fpath, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(v).strip() if v else '' for v in next(rows)]
        for r in rows:
            yield dict(zip(headers, [str(v).strip() if v else '' for v in r]))
    else:
        with open(fpath, 'r', encoding='utf-8', errors='ignore') as f:
            first = f.readline()
            delim = '\t' if '\t' in first else ','
            f.seek(0)
            reader = csv.DictReader(f, delimiter=delim)
            for row in reader:
                yield { (k or '').strip(): (v or '').strip() for k, v in row.items() }

def build_data(out_dir: str) -> dict:
    data = {}
    files_count = 0
    rows_total = 0
    rows_ok = 0

    for root, _, files in os.walk(out_dir):
        for fn in files:
            if not fn.lower().endswith(('.txt', '.csv', '.xlsx')):
                continue
            files_count += 1
            fpath = os.path.join(root, fn)
            for row in iter_rows_from_file(fpath):
                rows_total += 1
                npa_nxx = row.get('NPA-NXX') or row.get('NPA NXX') or ''
                npa, nxx = '', ''
                if '-' in npa_nxx:
                    a, b = npa_nxx.split('-', 1)
                    if a.isdigit() and b.isdigit():
                        npa, nxx = a, b
                elif len(npa_nxx) == 6 and npa_nxx.isdigit():
                    npa, nxx = npa_nxx[:3], npa_nxx[3:]
                else:
                    # try separate columns
                    a = (row.get('NPA') or '').strip()
                    b = (row.get('NXX') or '').strip()
                    if a.isdigit() and b.isdigit():
                        npa, nxx = a, b

                if not (npa and nxx and len(npa)==3 and len(nxx)==3 and npa.isdigit() and nxx.isdigit()):
                    continue

                prefix = f"{npa}{nxx}"
                company_original = row.get('Company') or row.get('Operating Company Name') or ''
                ocn = row.get('OCN') or row.get('OCN ') or ''
                rate_center = row.get('RateCenter') or row.get('Rate Center') or ''
                state_abbr = (row.get('State') or '').strip()
                state_full = STATE_MAP.get(state_abbr.upper(), state_abbr)
                city = safe_title(rate_center)

                # Parent-only carrier name
                carrier = normalize_brand(company_original) or company_original

                # Type detection (no guessing)
                _brand = normalize_brand(company_original)
                _type = detect_type(company_original, _brand)

                rec = {
                    'prefix': prefix,
                    'ocn': ocn,
                    'company': company_original,          # keep for compatibility
                    'company_original': company_original, # explicit original
                    'carrier': carrier,                   # parent-only branding if known
                    'type': _type,                        # may be blank if unknown
                    'rate_center': rate_center,
                    'city': city,
                    'state': state_full,
                    'last_source': fn
                }
                data[prefix] = rec
                rows_ok += 1

    log(f"Parsed files: {files_count}, rows read: {rows_total:,}, rows accepted: {rows_ok:,}")
    return data

# ---------------------------- SAVE ---------------------------------
def backup_json():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    if os.path.exists(OUTPUT_JSON):
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M')
        shutil.copy2(OUTPUT_JSON, os.path.join(BACKUP_DIR, f"data_{ts}.json"))
        log(f"Backup saved -> backups/data_{ts}.json")

def save_json(data: dict):
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    log(f"Wrote {OUTPUT_JSON} ({len(data):,} prefixes)")

def save_sqlite(data: dict):
    conn = sqlite3.connect(OUTPUT_DB)
    cur = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS prefixes")
    cur.execute(
        """
        CREATE TABLE prefixes (
            prefix TEXT PRIMARY KEY,
            ocn TEXT,
            company TEXT,
            company_original TEXT,
            carrier TEXT,
            type TEXT,
            rate_center TEXT,
            city TEXT,
            state TEXT,
            last_source TEXT
        )
        """
    )
    cur.executemany(
        "INSERT INTO prefixes (prefix, ocn, company, company_original, carrier, type, rate_center, city, state, last_source) VALUES (?,?,?,?,?,?,?,?,?,?)",
        [(
            rec['prefix'], rec['ocn'], rec['company'], rec['company_original'], rec['carrier'], rec['type'], rec['rate_center'], rec['city'], rec['state'], rec['last_source']
        ) for rec in data.values()]
    )
    conn.commit()
    conn.close()
    log(f"Wrote {OUTPUT_DB}")

# ---------------------------- SUMMARY ---------------------------------
def show_summary(data: dict):
    total = len(data)
    buckets = {"Mobile":0, "VoIP":0, "Paging":0, "Landline":0, "Unknown":0}
    for rec in data.values():
        t = rec.get('type') or ""
        if t in buckets:
            buckets[t] += 1
        else:
            buckets["Unknown"] += 1
    log("---- Update Summary ----")
    log(f"Total Prefixes: {total:,}")
    for k in ["Mobile","Landline","VoIP","Paging","Unknown"]:
        log(f"  {k}: {buckets[k]:,}")
    log("-------------------------")

def safe_title(s: str) -> str:
    return s.title().replace("_", " ") if s else ""

# ---------------------------- MAIN ---------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--check-only', action='store_true', help='Only list .zip links and Last-Modified (no download/parse)')
    ap.add_argument('--no-fetch', action='store_true', help='Skip downloading; process existing zips in nanpa_zips')
    args = ap.parse_args()

    session = make_session()

    if args.check_only:
        try:
            links = get_zip_links(session)
            log("Available ZIPs:")
            for url in links:
                try:
                    h = session.head(url, allow_redirects=True, timeout=20)
                    lm = h.headers.get('Last-Modified', 'Unknown')
                    sz = h.headers.get('Content-Length', 'Unknown')
                    print(f"- {url}  | Last-Modified: {lm} | Size: {sz}")
                except Exception as e:
                    print(f"- {url}  | (HEAD failed: {e})")
        except Exception as e:
            log(f"ERROR while checking: {e}")
        sys.exit(0)

    os.makedirs(ZIPS_DIR, exist_ok=True)

    if not args.no_fetch:
        new_files = fetch_latest(ZIPS_DIR, session)
        if not new_files:
            log("No new files downloaded (using existing zips)")

    extract_all(ZIPS_DIR, FILES_DIR)

    log("Building prefix dataset ...")
    data = build_data(FILES_DIR)

    backup_json()
    save_json(data)
    save_sqlite(data)
    show_summary(data)
    log("✅ Update completed successfully!")

if __name__ == '__main__':
    main()
